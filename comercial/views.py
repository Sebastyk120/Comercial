import io
from collections import defaultdict
from datetime import datetime
from decimal import Decimal
from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test, login_required
from django.db import transaction
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.views.generic import TemplateView
from django.views.generic.edit import CreateView, UpdateView
from django_tables2 import SingleTableView
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.workbook import Workbook
from .forms import SearchForm, PedidoForm, EditarPedidoForm, EliminarPedidoForm, DetallePedidoForm, \
    EliminarDetallePedidoForm, EditarPedidoExportadorForm, EditarDetallePedidoForm
from .models import Pedido, DetallePedido
from .resources import obtener_datos_con_totales, crear_archivo_excel, obtener_datos_con_totales_etnico, \
    obtener_datos_con_totales_fieldex, obtener_datos_con_totales_juan
from .tables import PedidoTable, DetallePedidoTable, PedidoExportadorTable, CarteraPedidoTable, ComisionPedidoTable


# from .resources import CarteraPedidoResource


# -----------Funcion para permisos por grupo ---------------------
def es_miembro_del_grupo(nombre_grupo):
    def es_miembro(user):
        return user.groups.filter(name=nombre_grupo).exists()

    return es_miembro


# ------------------ Exportacion de Comisiones Excel General --------------------------------------------------------
class ExportarComisionesView(TemplateView):
    template_name = 'export_comisiones_general.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Agrega contexto adicional aquí si es necesario
        return context


@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Heavens'), login_url='home'))
def exportar_comisiones_excel(request):
    # Crear un libro de trabajo de Excel
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Comisiones Totales General'
    font = Font(bold=True)
    fill = PatternFill(start_color="3ef983", end_color="3ef983", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="3580e0", end_color="3580e0", fill_type="solid")
    total_align = Alignment(horizontal="center")

    # Encabezados
    columns = ['Pedido', 'Cliente', 'Exportador', 'Fecha Pago Cliente', 'Valor Total Factura USD', 'Estado Factura',
               'Trm Monetizacion', 'Valor Comision USD', 'Valor Comision Pesos', 'Documento Cobro Comision',
               'Fecha Pago Comision', 'Diferencia O Abono', 'Estado Comision', 'Cobrar comision']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill
    # Crear un diccionario para almacenar los totales de comisiones por exportadora
    totales_por_comision_usd = defaultdict(Decimal)
    totales_no_cobrables_por_exportadora = defaultdict(Decimal)
    totales_cobrados_por_exportadora = defaultdict(Decimal)
    totales_por_cobrar_por_exportadora = defaultdict(Decimal)

    fecha_inicial_str = request.POST.get('fecha_inicial')
    fecha_final_str = request.POST.get('fecha_final')

    # Verificar si las fechas están vacías
    if fecha_inicial_str and fecha_final_str:
        # Convertir las cadenas de fecha en datetime
        fecha_inicial = datetime.strptime(fecha_inicial_str, '%Y-%m-%d')
        fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d')

        # Filtrar los pedidos por fecha_entrega dentro del rango
        queryset = Pedido.objects.filter(fecha_entrega__gte=fecha_inicial, fecha_entrega__lte=fecha_final)
    else:
        # Si las fechas están vacías, exportar todos los pedidos
        queryset = Pedido.objects.all()

    # Obtener los datos
    for pedido in queryset:
        if pedido.diferencia_por_abono < 0:
            totales_no_cobrables_por_exportadora[pedido.exportadora.nombre] += Decimal(pedido.valor_total_comision_usd)
        if pedido.fecha_pago_comision is not None:
            totales_cobrados_por_exportadora[pedido.exportadora.nombre] += Decimal(pedido.valor_total_comision_usd)
        if pedido.fecha_pago_comision is None and pedido.diferencia_por_abono >= 0:
            totales_por_cobrar_por_exportadora[pedido.exportadora.nombre] += Decimal(pedido.valor_total_comision_usd)
        totales_por_comision_usd[pedido.exportadora.nombre] += Decimal(pedido.valor_total_comision_usd)

    # Agregar datos al libro de trabajo
    for row_num, pedido in enumerate(queryset, start=2):
        cobrar_comision = "Sí" if pedido.diferencia_por_abono >= 0 else "No"
        row = [
            pedido.pk,
            pedido.cliente.nombre,
            pedido.exportadora.nombre,
            pedido.fecha_pago_comision,
            pedido.valor_total_factura_usd,
            pedido.estado_factura,
            pedido.trm_monetizacion,
            pedido.valor_total_comision_usd,
            pedido.valor_comision_pesos,
            pedido.documento_cobro_comision,
            pedido.fecha_pago_comision,
            pedido.diferencia_por_abono,
            pedido.estado_comision,
            cobrar_comision,  # Añadido valor de 'Cobrar comision'
        ]
        for col_num, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    # Agregar los totales al final de la hoja de trabajo

    def aplicar_estilo_total(fila):  # Funcion Para Dar Estilo A Los Totales.
        for col in range(1, len(columns) + 1):
            celda = worksheet.cell(row=fila, column=col)
            celda.font = total_font
            celda.fill = total_fill
            celda.alignment = total_align

    row_num += 2  # Saltar a la siguiente fila después de los datos
    for exportadora, total in totales_por_comision_usd.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Comisiónes USD " + exportadora)
        worksheet.cell(row=row_num, column=3, value=total)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila
    for exportadora, total_no_cobrable in totales_no_cobrables_por_exportadora.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Comisiónes No Cobrables USD " + exportadora)
        worksheet.cell(row=row_num, column=3, value=total_no_cobrable)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila
    for exportadora, total_cobrado in totales_cobrados_por_exportadora.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Comisiónes Cobradas " + exportadora)
        worksheet.cell(row=row_num, column=3, value=total_cobrado)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila
    for exportadora, total_por_cobrar in totales_por_cobrar_por_exportadora.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Por Cobrar " + exportadora)
        worksheet.cell(row=row_num, column=3, value=total_por_cobrar)
        aplicar_estilo_total(row_num)
        row_num += 1

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="comisiones_pedidos_general.xlsx"'

    return response


# ------------------ Exportacion de Comisiones Excel Etnico --------------------------------------------------------
@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Etnico'), login_url='home'))
def exportar_comisiones_etnico(request):
    # Crear un libro de trabajo de Excel
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Comisiones Totales Etnico'
    font = Font(bold=True)
    fill = PatternFill(start_color="3ef983", end_color="3ef983", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="3580e0", end_color="3580e0", fill_type="solid")
    total_align = Alignment(horizontal="center")

    # Encabezados
    columns = ['Pedido', 'Cliente', 'Exportador', 'Fecha Pago Cliente', 'Valor Total Factura USD', 'Estado Factura',
               'Trm Monetizacion', 'Valor Comision USD', 'Valor Comision Pesos', 'Documento Cobro Comision',
               'Fecha Pago Comision', 'Diferencia O Abono', 'Estado Comision', 'Cobrar comision']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill
    # Crear un diccionario para almacenar los totales de comisiones por exportadora
    totales_por_comision_usd = defaultdict(Decimal)
    totales_no_cobrables_por_exportadora = defaultdict(Decimal)
    totales_cobrados_por_exportadora = defaultdict(Decimal)
    totales_por_cobrar_por_exportadora = defaultdict(Decimal)

    # Obtener los datos de tu modelo y calcular los totales
    for pedido in Pedido.objects.filter(exportadora__nombre='Etnico'):
        if pedido.diferencia_por_abono < 0:
            totales_no_cobrables_por_exportadora[pedido.exportadora.nombre] += Decimal(pedido.valor_total_comision_usd)
        if pedido.fecha_pago_comision is not None:
            totales_cobrados_por_exportadora[pedido.exportadora.nombre] += Decimal(pedido.valor_total_comision_usd)
        if pedido.fecha_pago_comision is None and pedido.diferencia_por_abono >= 0:
            totales_por_cobrar_por_exportadora[pedido.exportadora.nombre] += Decimal(pedido.valor_total_comision_usd)
        totales_por_comision_usd[pedido.exportadora.nombre] += Decimal(pedido.valor_total_comision_usd)
    # Obtener los datos de tu modelo
    queryset = Pedido.objects.filter(exportadora__nombre='Etnico')

    # Agregar datos al libro de trabajo
    for row_num, pedido in enumerate(queryset, start=2):
        cobrar_comision = "Sí" if pedido.diferencia_por_abono >= 0 else "No"
        row = [
            pedido.pk,
            pedido.cliente.nombre,
            pedido.exportadora.nombre,
            pedido.fecha_pago,
            pedido.valor_total_factura_usd,
            pedido.estado_factura,
            pedido.trm_monetizacion,
            pedido.valor_total_comision_usd,
            pedido.valor_comision_pesos,
            pedido.documento_cobro_comision,
            pedido.fecha_pago_comision,
            pedido.diferencia_por_abono,
            pedido.estado_comision,
            cobrar_comision,  # Añadido valor de 'Cobrar comision'
        ]
        for col_num, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    # Agregar los totales al final de la hoja de trabajo

    """def aplicar_estilo_total(fila):  # Funcion APara Dar Estilo A Los Totales.
        for col in range(1, len(columns) + 1):
            celda = worksheet.cell(row=fila, column=col)
            celda.font = total_font
            celda.fill = total_fill
            celda.alignment = total_align

    row_num += 2  # Saltar a la siguiente fila después de los datos
    for exportadora, total in totales_por_comision_usd.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Comisiónes USD")
        worksheet.cell(row=row_num, column=3, value=total)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila
    for exportadora, total_no_cobrable in totales_no_cobrables_por_exportadora.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Comisiónes No Cobrables USD")
        worksheet.cell(row=row_num, column=3, value=total_no_cobrable)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila
    for exportadora, total_cobrado in totales_cobrados_por_exportadora.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Comisiónes Pagadas")
        worksheet.cell(row=row_num, column=3, value=total_cobrado)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila
    for exportadora, total_por_cobrar in totales_por_cobrar_por_exportadora.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Comisiones Por Pagar")
        worksheet.cell(row=row_num, column=3, value=total_por_cobrar)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila"""

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="comisiones_pedidos_etnico.xlsx"'

    return response


# ------------------ Exportacion de Comisiones Excel Fieldex --------------------------------------------------------
@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Fieldex'), login_url='home'))
def exportar_comisiones_fieldex(request):
    # Crear un libro de trabajo de Excel
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Comisiones Totales Fieldex'
    font = Font(bold=True)
    fill = PatternFill(start_color="3ef983", end_color="3ef983", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="3580e0", end_color="3580e0", fill_type="solid")
    total_align = Alignment(horizontal="center")

    # Encabezados
    columns = ['Pedido', 'Cliente', 'Exportador', 'Fecha Pago Cliente', 'Valor Total Factura USD', 'Estado Factura',
               'Trm Monetizacion', 'Valor Comision USD', 'Valor Comision Pesos', 'Documento Cobro Comision',
               'Fecha Pago Comision', 'Diferencia O Abono', 'Estado Comision', 'Cobrar comision']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill
    # Crear un diccionario para almacenar los totales de comisiones por exportadora
    totales_por_comision_usd = defaultdict(Decimal)
    totales_no_cobrables_por_exportadora = defaultdict(Decimal)
    totales_cobrados_por_exportadora = defaultdict(Decimal)
    totales_por_cobrar_por_exportadora = defaultdict(Decimal)

    # Obtener los datos de tu modelo y calcular los totales
    for pedido in Pedido.objects.filter(exportadora__nombre='Fieldex'):
        if pedido.diferencia_por_abono < 0:
            totales_no_cobrables_por_exportadora[pedido.exportadora.nombre] += Decimal(pedido.valor_total_comision_usd)
        if pedido.fecha_pago_comision is not None:
            totales_cobrados_por_exportadora[pedido.exportadora.nombre] += Decimal(pedido.valor_total_comision_usd)
        if pedido.fecha_pago_comision is None and pedido.diferencia_por_abono > 0:
            totales_por_cobrar_por_exportadora[pedido.exportadora.nombre] += Decimal(pedido.valor_total_comision_usd)
        totales_por_comision_usd[pedido.exportadora.nombre] += Decimal(pedido.valor_total_comision_usd)
    # Obtener los datos de tu modelo
    queryset = Pedido.objects.filter(exportadora__nombre='Fieldex')

    # Agregar datos al libro de trabajo
    for row_num, pedido in enumerate(queryset, start=2):
        cobrar_comision = "Sí" if pedido.diferencia_por_abono >= 0 else "No"
        row = [
            pedido.pk,
            pedido.cliente.nombre,
            pedido.exportadora.nombre,
            pedido.fecha_pago,
            pedido.valor_total_factura_usd,
            pedido.estado_factura,
            pedido.trm_monetizacion,
            pedido.valor_total_comision_usd,
            pedido.valor_comision_pesos,
            pedido.documento_cobro_comision,
            pedido.fecha_pago_comision,
            pedido.diferencia_por_abono,
            pedido.estado_comision,
            cobrar_comision,  # Añadido valor de 'Cobrar comision'
        ]
        for col_num, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    # Agregar los totales al final de la hoja de trabajo

    """def aplicar_estilo_total(fila):  # Funcion APara Dar Estilo A Los Totales.
        for col in range(1, len(columns) + 1):
            celda = worksheet.cell(row=fila, column=col)
            celda.font = total_font
            celda.fill = total_fill
            celda.alignment = total_align

    row_num += 2  # Saltar a la siguiente fila después de los datos
    for exportadora, total in totales_por_comision_usd.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Comisiónes USD")
        worksheet.cell(row=row_num, column=3, value=total)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila
    for exportadora, total_no_cobrable in totales_no_cobrables_por_exportadora.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Comisiónes No Cobrables USD")
        worksheet.cell(row=row_num, column=3, value=total_no_cobrable)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila
    for exportadora, total_cobrado in totales_cobrados_por_exportadora.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Comisiónes Pagadas")
        worksheet.cell(row=row_num, column=3, value=total_cobrado)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila
    for exportadora, total_por_cobrar in totales_por_cobrar_por_exportadora.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Comisiones Por Pagar")
        worksheet.cell(row=row_num, column=3, value=total_por_cobrar)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila"""

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="comisiones_pedidos_fieldex.xlsx"'

    return response


# ------------------ Exportacion de Comisiones Excel Juan Matas --------------------------------------------------------
@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Juan_Matas'), login_url='home'))
def exportar_comisiones_juan(request):
    # Crear un libro de trabajo de Excel
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Comisiones Totales Juan Matas'
    font = Font(bold=True)
    fill = PatternFill(start_color="3ef983", end_color="3ef983", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="3580e0", end_color="3580e0", fill_type="solid")
    total_align = Alignment(horizontal="center")

    # Encabezados
    columns = ['Pedido', 'Cliente', 'Exportador', 'Fecha Pago Cliente', 'Valor Total Factura USD', 'Estado Factura',
               'Trm Monetizacion', 'Valor Comision USD', 'Valor Comision Pesos', 'Documento Cobro Comision',
               'Fecha Pago Comision', 'Diferencia O Abono', 'Estado Comision', 'Cobrar comision']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill
    # Crear un diccionario para almacenar los totales de comisiones por exportadora
    totales_por_comision_usd = defaultdict(Decimal)
    totales_no_cobrables_por_exportadora = defaultdict(Decimal)
    totales_cobrados_por_exportadora = defaultdict(Decimal)
    totales_por_cobrar_por_exportadora = defaultdict(Decimal)

    # Obtener los datos de tu modelo y calcular los totales
    for pedido in Pedido.objects.filter(exportadora__nombre='Juan_Matas'):
        if pedido.diferencia_por_abono < 0:
            totales_no_cobrables_por_exportadora[pedido.exportadora.nombre] += Decimal(pedido.valor_total_comision_usd)
        if pedido.fecha_pago_comision is not None:
            totales_cobrados_por_exportadora[pedido.exportadora.nombre] += Decimal(pedido.valor_total_comision_usd)
        if pedido.fecha_pago_comision is None and pedido.diferencia_por_abono > 0:
            totales_por_cobrar_por_exportadora[pedido.exportadora.nombre] += Decimal(pedido.valor_total_comision_usd)
        totales_por_comision_usd[pedido.exportadora.nombre] += Decimal(pedido.valor_total_comision_usd)
    # Obtener los datos de tu modelo
    queryset = Pedido.objects.filter(exportadora__nombre='Juan_Matas')

    # Agregar datos al libro de trabajo
    for row_num, pedido in enumerate(queryset, start=2):
        cobrar_comision = "Sí" if pedido.diferencia_por_abono >= 0 else "No"
        row = [
            pedido.pk,
            pedido.cliente.nombre,
            pedido.exportadora.nombre,
            pedido.fecha_pago,
            pedido.valor_total_factura_usd,
            pedido.estado_factura,
            pedido.trm_monetizacion,
            pedido.valor_total_comision_usd,
            pedido.valor_comision_pesos,
            pedido.documento_cobro_comision,
            pedido.fecha_pago_comision,
            pedido.diferencia_por_abono,
            pedido.estado_comision,
            cobrar_comision,  # Añadido valor de 'Cobrar comision'
        ]
        for col_num, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    # Agregar los totales al final de la hoja de trabajo

    """def aplicar_estilo_total(fila):  # Funcion APara Dar Estilo A Los Totales.
        for col in range(1, len(columns) + 1):
            celda = worksheet.cell(row=fila, column=col)
            celda.font = total_font
            celda.fill = total_fill
            celda.alignment = total_align

    row_num += 2  # Saltar a la siguiente fila después de los datos
    for exportadora, total in totales_por_comision_usd.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Comisiónes USD")
        worksheet.cell(row=row_num, column=3, value=total)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila
    for exportadora, total_no_cobrable in totales_no_cobrables_por_exportadora.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Comisiónes No Cobrables USD")
        worksheet.cell(row=row_num, column=3, value=total_no_cobrable)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila
    for exportadora, total_cobrado in totales_cobrados_por_exportadora.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Comisiónes Pagadas")
        worksheet.cell(row=row_num, column=3, value=total_cobrado)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila
    for exportadora, total_por_cobrar in totales_por_cobrar_por_exportadora.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Comisiones Por Pagar")
        worksheet.cell(row=row_num, column=3, value=total_por_cobrar)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila"""

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="comisiones_pedidos_juan.xlsx"'

    return response


# ------------------ Exportacion de Detalles de Pedidos Excel General -------------------------------------
@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Heavens'), login_url='home'))
def exportar_detalles_pedidos_excel(request):
    # Crear un libro de trabajo de Excel
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Detalles De Pedidos General'
    font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color="251819", end_color="251819", fill_type="solid")

    # Encabezados
    columns = ['Pedido', 'Cliente', 'Fruta', 'Presentacion', 'Cajas Solicitadas', 'Peso Presentacion', 'kilos',
               'Cajas Enviadas',
               'Kilos Enviados', 'Diferencia', 'Tipo Caja', 'Referencia', 'Stiker', 'Lleva Contenedor',
               'Ref Contenedor', 'Cant Contenedor', 'Tarifa Comision', 'Valor x Caja USD', 'Valor X Producto',
               'No Cajas NC', 'Valor NC', 'Afecta Comision', 'Valor Total Comision Producto']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill

    # Obtener los datos de tu modelo
    queryset = DetallePedido.objects.all()

    # Agregar datos al libro de trabajo
    for row_num, detalle in enumerate(queryset, start=2):
        row = [
            detalle.pedido.pk,
            detalle.pedido.cliente.nombre,
            detalle.fruta.nombre,
            detalle.presentacion.nombre,
            detalle.cajas_solicitadas,
            detalle.presentacion_peso,
            detalle.kilos,
            detalle.cajas_enviadas,
            detalle.kilos_enviados,
            detalle.diferencia,
            detalle.tipo_caja.nombre,
            detalle.referencia.nombre,
            detalle.stickers,
            detalle.lleva_contenedor,
            detalle.referencia_contenedor,
            detalle.cantidad_contenedores,
            detalle.tarifa_comision,
            detalle.valor_x_caja_usd,
            detalle.valor_x_producto,
            detalle.no_cajas_nc,
            detalle.valor_nota_credito_usd,
            detalle.afecta_comision,
            detalle.valor_total_comision_x_producto,
        ]
        for col_num, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Detalles_Pedidos.xlsx"'

    return response


# ------------------ Exportacion de Pedidos Excel General --------------------------------------------------------

# Vista Para Exportar Pedidos:

class ExportarPedidosView(TemplateView):
    template_name = 'export_pedidos_general.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Agrega contexto adicional aquí si es necesario
        return context


# Exportacion de Pedidos OpenPyXl >
@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Heavens'), login_url='home'))
def exportar_pedidos_excel(request):
    # Crear un libro de trabajo de Excel
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Pedidos Totales General'
    font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color="1e0c42", end_color="1e0c42", fill_type="solid")

    # Encabezados
    columns = ['No', 'Cliente', 'Fecha Solicitud', 'Fecha Entrega', 'Exportador', 'Dias Cartera', 'awb',
               'Destino', 'Numero Factura', 'Total Cajas Enviadas', 'No NC', 'Motivo NC', 'Valor Total NC',
               'Tasa Representatativa', 'Valor Pagado Cliente', 'Comision Bancaria USD', 'Fecha Pago',
               'TRM Monetizacion', 'Estado Factura', 'Diferencia Pago', 'Dias Vencimiento', 'Valor Total Factura USD',
               'Valor Comision USD', 'Valor Comision Pesos', 'Documento Cobro Comision', 'Fecha Pago Comision',
               'Estado Comision']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill

    fecha_inicial_str = request.POST.get('fecha_inicial')
    fecha_final_str = request.POST.get('fecha_final')

    # Verificar si las fechas están vacías o nulas
    if fecha_inicial_str and fecha_final_str:
        # Convertir las cadenas de fecha en objetos datetime
        fecha_inicial = datetime.strptime(fecha_inicial_str, '%Y-%m-%d')
        fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d')

        # Filtrar los pedidos por fecha_entrega dentro del rango
        queryset = Pedido.objects.filter(fecha_entrega__gte=fecha_inicial, fecha_entrega__lte=fecha_final)
    else:
        # Si las fechas están vacías, exportar todos los pedidos
        queryset = Pedido.objects.all()

    # Agregar datos al libro de trabajo
    for row_num, pedido in enumerate(queryset, start=2):
        row = [
            pedido.pk,
            pedido.cliente.nombre,
            pedido.fecha_solicitud,
            pedido.fecha_entrega,
            pedido.exportadora.nombre,
            pedido.dias_cartera,
            pedido.awb,
            pedido.destino,
            pedido.numero_factura,
            pedido.total_cajas_enviadas,
            pedido.nota_credito_no,
            pedido.motivo_nota_credito,
            pedido.valor_total_nota_credito_usd,
            pedido.tasa_representativa_usd_diaria,
            pedido.valor_pagado_cliente_usd,
            pedido.comision_bancaria_usd,
            pedido.fecha_pago,
            pedido.trm_monetizacion,
            pedido.estado_factura,
            pedido.diferencia_por_abono,
            pedido.dias_de_vencimiento,
            pedido.valor_total_factura_usd,
            pedido.valor_total_comision_usd,
            pedido.valor_comision_pesos,
            pedido.documento_cobro_comision,
            pedido.fecha_pago_comision,
            pedido.estado_comision,
        ]
        for col_num, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="pedido_general.xlsx"'

    return response


# ------------------ Exportacion de Pedidos Excel Etnico --------------------------------------------------------

# Vista Para Exportar Pedidos Etnico:

class ExportarPedidosEtnicoView(TemplateView):
    template_name = 'export_pedidos_etnico.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Agrega contexto adicional aquí si es necesario
        return context


@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Etnico'), login_url='home'))
def exportar_pedidos_etnico(request):
    # Crear un libro de trabajo de Excel
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Pedidos Totales Etnico'
    font = Font(bold=True)
    fill = PatternFill(start_color="fffaac", end_color="fffaac", fill_type="solid")

    # Encabezados
    columns = ['Cliente', 'Fecha Solicitud', 'Fecha Entrega', 'Exportador', 'Dias Cartera', 'awb',
               'Destino', 'Numero Factura', 'Total Cajas Enviadas', 'No NC', 'Motivo NC', 'Valor Total NC',
               'Tasa Representatativa', 'Valor Pagado Cliente', 'Comision Bancaria USD', 'Fecha Pago',
               'TRM Monetizacion', 'Estado Factura', 'Diferencia Pago', 'Dias Vencimiento', 'Valor Total Factura USD',
               'Valor Comision USD', 'Valor Comision Pesos', 'Documento Cobro Comision', 'Fecha Pago Comision',
               'Estado Comision']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill

    fecha_inicial_str = request.POST.get('fecha_inicial')
    fecha_final_str = request.POST.get('fecha_final')

    # Verificar si las fechas están vacías o nulas
    if fecha_inicial_str and fecha_final_str:
        # Convertir las cadenas de fecha en objetos datetime
        fecha_inicial = datetime.strptime(fecha_inicial_str, '%Y-%m-%d')
        fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d')

        # Filtrar los pedidos por fecha_entrega dentro del rango
        queryset = Pedido.objects.filter(exportadora__nombre='Etnico', fecha_entrega__gte=fecha_inicial,
                                         fecha_entrega__lte=fecha_final)
    else:
        # Si las fechas están vacías, exportar todos los pedidos de la exportadora 'Etnico'
        queryset = Pedido.objects.filter(exportadora__nombre='Etnico')

    # Agregar datos al libro de trabajo
    for row_num, pedido in enumerate(queryset, start=2):
        row = [
            pedido.cliente.nombre,
            pedido.fecha_solicitud,
            pedido.fecha_entrega,
            pedido.exportadora.nombre,
            pedido.dias_cartera,
            pedido.awb,
            pedido.destino,
            pedido.numero_factura,
            pedido.total_cajas_enviadas,
            pedido.nota_credito_no,
            pedido.motivo_nota_credito,
            pedido.valor_total_nota_credito_usd,
            pedido.tasa_representativa_usd_diaria,
            pedido.valor_pagado_cliente_usd,
            pedido.comision_bancaria_usd,
            pedido.fecha_pago,
            pedido.trm_monetizacion,
            pedido.estado_factura,
            pedido.diferencia_por_abono,
            pedido.dias_de_vencimiento,
            pedido.valor_total_factura_usd,
            pedido.valor_total_comision_usd,
            pedido.valor_comision_pesos,
            pedido.documento_cobro_comision,
            pedido.fecha_pago_comision,
            pedido.estado_comision,
        ]
        for col_num, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="pedidos_etnico.xlsx"'

    return response


# ------------------ Exportacion de Pedidos Excel Fieldex --------------------------------------------------------

class ExportarPedidosFieldexView(TemplateView):
    template_name = 'export_pedidos_fieldex.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Agrega contexto adicional aquí si es necesario
        return context


@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Fieldex'), login_url='home'))
def exportar_pedidos_fieldex(request):
    # Crear un libro de trabajo de Excel
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Pedidos Totales Fieldex'
    font = Font(bold=True)
    fill = PatternFill(start_color="fffaac", end_color="fffaac", fill_type="solid")

    # Encabezados
    columns = ['Cliente', 'Fecha Solicitud', 'Fecha Entrega', 'Exportador', 'Dias Cartera', 'awb',
               'Destino', 'Numero Factura', 'Total Cajas Enviadas', 'No NC', 'Motivo NC', 'Valor Total NC',
               'Tasa Representatativa', 'Valor Pagado Cliente', 'Comision Bancaria USD', 'Fecha Pago',
               'TRM Monetizacion', 'Estado Factura', 'Diferencia Pago', 'Dias Vencimiento', 'Valor Total Factura USD',
               'Valor Comision USD', 'Valor Comision Pesos', 'Documento Cobro Comision', 'Fecha Pago Comision',
               'Estado Comision']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill

    fecha_inicial_str = request.POST.get('fecha_inicial')
    fecha_final_str = request.POST.get('fecha_final')

    # Verificar si las fechas están vacías o nulas
    if fecha_inicial_str and fecha_final_str:
        # Convertir las cadenas de fecha en objetos datetime
        fecha_inicial = datetime.strptime(fecha_inicial_str, '%Y-%m-%d')
        fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d')

        # Filtrar los pedidos por fecha_entrega dentro del rango
        queryset = Pedido.objects.filter(exportadora__nombre='Fieldex', fecha_entrega__gte=fecha_inicial,
                                         fecha_entrega__lte=fecha_final)
    else:
        # Si las fechas están vacías, exportar todos los pedidos de la exportadora 'Etnico'
        queryset = Pedido.objects.filter(exportadora__nombre='Fieldex')

    # Agregar datos al libro de trabajo
    for row_num, pedido in enumerate(queryset, start=2):
        row = [
            pedido.cliente.nombre,
            pedido.fecha_solicitud,
            pedido.fecha_entrega,
            pedido.exportadora.nombre,
            pedido.dias_cartera,
            pedido.awb,
            pedido.destino,
            pedido.numero_factura,
            pedido.total_cajas_enviadas,
            pedido.nota_credito_no,
            pedido.motivo_nota_credito,
            pedido.valor_total_nota_credito_usd,
            pedido.tasa_representativa_usd_diaria,
            pedido.valor_pagado_cliente_usd,
            pedido.comision_bancaria_usd,
            pedido.fecha_pago,
            pedido.trm_monetizacion,
            pedido.estado_factura,
            pedido.diferencia_por_abono,
            pedido.dias_de_vencimiento,
            pedido.valor_total_factura_usd,
            pedido.valor_total_comision_usd,
            pedido.valor_comision_pesos,
            pedido.documento_cobro_comision,
            pedido.fecha_pago_comision,
            pedido.estado_comision,
        ]
        for col_num, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="pedidos_fieldex.xlsx"'

    return response


# ------------------ Exportacion de Pedidos Excel Juan Matas ---------------------------------------------------------
class ExportarPedidosJuanView(TemplateView):
    template_name = 'export_pedidos_juan.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Agrega contexto adicional aquí si es necesario
        return context


@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Juan_Matas'), login_url='home'))
def exportar_pedidos_juan(request):
    # Crear un libro de trabajo de Excel
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Pedidos Totales Juan_Matas'
    font = Font(bold=True)
    fill = PatternFill(start_color="fffaac", end_color="fffaac", fill_type="solid")

    # Encabezados
    columns = ['Cliente', 'Fecha Solicitud', 'Fecha Entrega', 'Exportador', 'Dias Cartera', 'awb',
               'Destino', 'Numero Factura', 'Total Cajas Enviadas', 'No NC', 'Motivo NC', 'Valor Total NC',
               'Tasa Representatativa', 'Valor Pagado Cliente', 'Comision Bancaria USD', 'Fecha Pago',
               'TRM Monetizacion', 'Estado Factura', 'Diferencia Pago', 'Dias Vencimiento', 'Valor Total Factura USD',
               'Valor Comision USD', 'Valor Comision Pesos', 'Documento Cobro Comision', 'Fecha Pago Comision',
               'Estado Comision']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill

    fecha_inicial_str = request.POST.get('fecha_inicial')
    fecha_final_str = request.POST.get('fecha_final')

    # Verificar si las fechas están vacías o nulas
    if fecha_inicial_str and fecha_final_str:
        # Convertir las cadenas de fecha en objetos datetime
        fecha_inicial = datetime.strptime(fecha_inicial_str, '%Y-%m-%d')
        fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d')

        # Filtrar los pedidos por fecha_entrega dentro del rango
        queryset = Pedido.objects.filter(exportadora__nombre='Juan_Matas', fecha_entrega__gte=fecha_inicial,
                                         fecha_entrega__lte=fecha_final)
    else:
        # Si las fechas están vacías, exportar todos los pedidos de la exportadora 'Etnico'
        queryset = Pedido.objects.filter(exportadora__nombre='Juan_Matas')

    # Agregar datos al libro de trabajo
    for row_num, pedido in enumerate(queryset, start=2):
        row = [
            pedido.cliente.nombre,
            pedido.fecha_solicitud,
            pedido.fecha_entrega,
            pedido.exportadora.nombre,
            pedido.dias_cartera,
            pedido.awb,
            pedido.destino,
            pedido.numero_factura,
            pedido.total_cajas_enviadas,
            pedido.nota_credito_no,
            pedido.motivo_nota_credito,
            pedido.valor_total_nota_credito_usd,
            pedido.tasa_representativa_usd_diaria,
            pedido.valor_pagado_cliente_usd,
            pedido.comision_bancaria_usd,
            pedido.fecha_pago,
            pedido.trm_monetizacion,
            pedido.estado_factura,
            pedido.diferencia_por_abono,
            pedido.dias_de_vencimiento,
            pedido.valor_total_factura_usd,
            pedido.valor_total_comision_usd,
            pedido.valor_comision_pesos,
            pedido.documento_cobro_comision,
            pedido.fecha_pago_comision,
            pedido.estado_comision,
        ]
        for col_num, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="pedidos_juan.xlsx"'

    return response


# -------------------------- Funciones De Exportacion Cartera General--------------------------------------------------
@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Heavens'), login_url='home'))
def export_cartera_clientes(request):
    # Obtener los datos y los totales
    pedidos, totales = obtener_datos_con_totales()

    # Crear el archivo Excel
    ruta_archivo = 'estado_cuenta_clientes.xlsx'
    crear_archivo_excel(pedidos, totales, ruta_archivo)

    # Leer el archivo y preparar la respuesta
    with open(ruta_archivo, 'rb') as archivo_excel:
        response = HttpResponse(archivo_excel.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="estado_cuenta_clientes.xlsx"'

    return response


# -------------------------- Exportacion Cartera Etnico -------------------------------------------------------
@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Etnico'), login_url='home'))
def export_cartera_etnico(request):
    # Obtener los datos y los totales
    pedidos, totales = obtener_datos_con_totales_etnico()

    # Crear el archivo Excel
    ruta_archivo = 'etnico_cuenta_clientes.xlsx'
    crear_archivo_excel(pedidos, totales, ruta_archivo)

    # Leer el archivo y preparar la respuesta
    with open(ruta_archivo, 'rb') as archivo_excel:
        response = HttpResponse(archivo_excel.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="etnico_cuenta_clientes.xlsx"'

    return response


# -------------------------- Exportacion Cartera Fieldex -------------------------------------------------------
@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Fieldex'), login_url='home'))
def export_cartera_fieldex(request):
    # Obtener los datos y los totales
    pedidos, totales = obtener_datos_con_totales_fieldex()

    # Crear el archivo Excel
    ruta_archivo = 'fieldex_cuenta_clientes.xlsx'
    crear_archivo_excel(pedidos, totales, ruta_archivo)

    # Leer el archivo y preparar la respuesta
    with open(ruta_archivo, 'rb') as archivo_excel:
        response = HttpResponse(archivo_excel.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="fieldex_cuenta_clientes.xlsx"'

    return response


# -------------------------- Exportacion Cartera Juan Matas -------------------------------------------------------
@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Juan_Matas'), login_url='home'))
def export_cartera_juan(request):
    # Obtener los datos y los totales
    pedidos, totales = obtener_datos_con_totales_juan()

    # Crear el archivo Excel
    ruta_archivo = 'juan_cuenta_clientes.xlsx'
    crear_archivo_excel(pedidos, totales, ruta_archivo)

    # Leer el archivo y preparar la respuesta
    with open(ruta_archivo, 'rb') as archivo_excel:
        response = HttpResponse(archivo_excel.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="juan_cuenta_clientes.xlsx"'

    return response


# -------------------------------- Tabla De Pedidos General  ----------------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoListView(SingleTableView):
    model = Pedido
    table_class = PedidoTable
    table_pagination = {"per_page": 14}
    template_name = 'pedido_list_general.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset()
        form = self.form_class(self.request.GET)

        if form.is_valid() and form.cleaned_data.get('item_busqueda'):
            item_busqueda = form.cleaned_data.get('item_busqueda')

            # Aquí se maneja la lógica de la búsqueda
            try:
                item_busqueda_id = int(item_busqueda)  # Intenta convertir a entero
                queryset = queryset.filter(Q(cliente__nombre__icontains=item_busqueda) | Q(id=item_busqueda_id))
            except ValueError:
                # Si la conversión falla, busca solo por nombre
                queryset = queryset.filter(cliente__nombre__icontains=item_busqueda)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context


# -------------------------------- Tabla De Pedidos Etnico  ----------------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Etnico'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoEtnicoListView(SingleTableView):
    model = Pedido
    table_class = PedidoExportadorTable
    table_pagination = {"per_page": 14}
    template_name = 'pedido_list_etnico.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Etnico')
        form = self.form_class(self.request.GET)

        if form.is_valid() and form.cleaned_data.get('item_busqueda'):
            item_busqueda = form.cleaned_data.get('item_busqueda')

            # Aquí se maneja la lógica de la búsqueda
            try:
                item_busqueda_id = int(item_busqueda)  # Intenta convertir a entero
                queryset = queryset.filter(Q(cliente__nombre__icontains=item_busqueda) | Q(id=item_busqueda_id))
            except ValueError:
                # Si la conversión falla, busca solo por nombre
                queryset = queryset.filter(cliente__nombre__icontains=item_busqueda)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context


# -------------------------------- Tabla De Pedidos Fieldex  ----------------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Fieldex'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoFieldexListView(SingleTableView):
    model = Pedido
    table_class = PedidoExportadorTable
    table_pagination = {"per_page": 14}
    template_name = 'pedido_list_fieldex.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Fieldex')
        form = self.form_class(self.request.GET)

        if form.is_valid() and form.cleaned_data.get('item_busqueda'):
            item_busqueda = form.cleaned_data.get('item_busqueda')

            # Aquí se maneja la lógica de la búsqueda
            try:
                item_busqueda_id = int(item_busqueda)  # Intenta convertir a entero
                queryset = queryset.filter(Q(cliente__nombre__icontains=item_busqueda) | Q(id=item_busqueda_id))
            except ValueError:
                # Si la conversión falla, busca solo por nombre
                queryset = queryset.filter(cliente__nombre__icontains=item_busqueda)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context


# -------------------------------- Tabla De Pedidos Juan Matas  ----------------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Juan_Matas'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoJuanListView(SingleTableView):
    model = Pedido
    table_class = PedidoExportadorTable
    table_pagination = {"per_page": 14}
    template_name = 'pedido_list_Juan.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Juan_Matas')
        form = self.form_class(self.request.GET)

        if form.is_valid() and form.cleaned_data.get('item_busqueda'):
            item_busqueda = form.cleaned_data.get('item_busqueda')

            # Aquí se maneja la lógica de la búsqueda
            try:
                item_busqueda_id = int(item_busqueda)  # Intenta convertir a entero
                queryset = queryset.filter(Q(cliente__nombre__icontains=item_busqueda) | Q(id=item_busqueda_id))
            except ValueError:
                # Si la conversión falla, busca solo por nombre
                queryset = queryset.filter(cliente__nombre__icontains=item_busqueda)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context


# ----------------------------------- Mostrar Detalles De Pedido General ------------------------------------------
@method_decorator(login_required, name='dispatch')
class DetallePedidoListView(SingleTableView):
    model = DetallePedido
    table_class = DetallePedidoTable
    template_name = 'pedido_detalle_list.html'

    def get_queryset(self):
        pedido_id = self.kwargs.get('pedido_id')
        queryset = DetallePedido.objects.filter(pedido__id=pedido_id)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pedido_id = self.kwargs.get('pedido_id')
        context['pedido'] = get_object_or_404(Pedido, pk=pedido_id)  # Obtiene el objeto Pedido
        return context


# -------------------------------  Formulario - Crear Pedido General - Modal (General) ----------------------------
@method_decorator(login_required, name='dispatch')
class PedidoCreateView(CreateView):
    model = Pedido
    form_class = PedidoForm
    template_name = 'pedido_crear.html'
    success_url = '/pedido_list_general/'

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        return form

    @transaction.atomic
    def form_valid(self, form):
        self.object = form.save()  # Aquí es normal asignar a self.object
        messages.success(self.request,
                         f'El pedido para el cliente {form.cleaned_data['cliente']} se ha creado exitosamente.')
        return JsonResponse({'success': True})

    def form_invalid(self, form):
        return JsonResponse({'success': False, 'html': render_to_string(self.template_name, {'form': form})})


# -------------------------------  Formulario - Editar Pedido General - Modal (General) ----------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoUpdateView(UpdateView):
    model = Pedido
    form_class = EditarPedidoForm
    template_name = 'pedido_editar.html'
    success_url = '/pedido_list_general/'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.object = None

    def get_object(self, queryset=None):
        pedido_id = self.request.POST.get('pedido_id')
        pedido = get_object_or_404(Pedido, id=pedido_id)
        return pedido

    def get(self, request, *args, **kwargs):
        pedido_id = request.GET.get('pedido_id')
        self.object = get_object_or_404(Pedido, id=pedido_id)
        formatted_fecha_solicitud = self.object.fecha_solicitud.strftime('%Y-%m-%d')
        formatted_fecha_entrega = self.object.fecha_entrega.strftime('%Y-%m-%d')
        if self.object.fecha_pago_comision is None:
            form = self.form_class(
                instance=self.object,
                initial={'fecha_solicitud': formatted_fecha_solicitud, 'fecha_entrega': formatted_fecha_entrega}
            )
        else:
            formatted_fecha_pago_comision = self.object.fecha_pago_comision.strftime('%Y-%m-%d')
            form = self.form_class(
                instance=self.object,
                initial={'fecha_solicitud': formatted_fecha_solicitud, 'fecha_entrega': formatted_fecha_entrega,
                         'fecha_pago_comision': formatted_fecha_pago_comision}
            )
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, {'form': form}, request=request)
            return JsonResponse({'form': form_html})
        else:
            return super().get(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['instance'] = self.object
        return kwargs

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request,
                         f'El pedido para el cliente {form.cleaned_data['cliente']} se ha editado exitosamente.')
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse(
                {'success': False, 'html': render_to_string(self.template_name, {'form': form}, request=self.request)})
        else:
            return super().form_invalid(form)


# -------------------------------  //// Formulario - Editar Pedido Por Exportador //// ----------------------------
@method_decorator(login_required, name='dispatch')
class PedidoExportadorUpdateView(UpdateView):
    model = Pedido
    form_class = EditarPedidoExportadorForm
    template_name = 'pedido_editar.html'
    success_url = '/update_items/'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.object = None

    def get_object(self, queryset=None):
        pedido_id = self.request.POST.get('pedido_id')
        pedido = get_object_or_404(Pedido, id=pedido_id)
        return pedido

    def get(self, request, *args, **kwargs):
        pedido_id = request.GET.get('pedido_id')
        self.object = get_object_or_404(Pedido, id=pedido_id)
        formatted_fecha_solicitud = self.object.fecha_solicitud.strftime('%Y-%m-%d')
        formatted_fecha_entrega = self.object.fecha_entrega.strftime('%Y-%m-%d')
        if self.object.fecha_pago_comision is None:
            form = self.form_class(
                instance=self.object,
                initial={'fecha_solicitud': formatted_fecha_solicitud, 'fecha_entrega': formatted_fecha_entrega}
            )
        else:
            formatted_fecha_pago_comision = self.object.fecha_pago_comision.strftime('%Y-%m-%d')
            form = self.form_class(
                instance=self.object,
                initial={'fecha_solicitud': formatted_fecha_solicitud, 'fecha_entrega': formatted_fecha_entrega,
                         'fecha_pago_comision': formatted_fecha_pago_comision}
            )
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, {'form': form}, request=request)
            return JsonResponse({'form': form_html})
        else:
            return super().get(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['instance'] = self.object
        return kwargs

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request,
                         f'El pedido se ha editado exitosamente.')
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse(
                {'success': False, 'html': render_to_string(self.template_name, {'form': form}, request=self.request)})
        else:
            return super().form_invalid(form)


# -------------------------------  Formulario - Eliminar Pedido General - Modal (General) ----------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoDeleteView(UpdateView):
    model = Pedido
    form_class = EliminarPedidoForm
    template_name = 'pedido_eliminar.html'
    success_url = '/pedido_list_general/'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.object = None

    def get_object(self, queryset=None):
        pedido_id = self.request.POST.get('pedido_id')
        pedido = get_object_or_404(Pedido, id=pedido_id)
        return pedido

    def get(self, request, *args, **kwargs):
        pedido_id = request.GET.get('pedido_id')
        self.object = get_object_or_404(Pedido, id=pedido_id)
        formatted_fecha_solicitud = self.object.fecha_solicitud.strftime('%Y-%m-%d')
        formatted_fecha_entrega = self.object.fecha_entrega.strftime('%Y-%m-%d')
        form = self.form_class(
            instance=self.object,
            initial={'fecha_solicitud': formatted_fecha_solicitud, 'fecha_entrega': formatted_fecha_entrega}
        )
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, {'form': form}, request=request)
            return JsonResponse({'form': form_html})
        else:
            return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        pedido = form.save(commit=False)
        pedido.delete()
        messages.success(self.request,
                         f'El pedido para el cliente {form.cleaned_data['cliente']} se ha eliminado exitosamente.')
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse(
                {'success': False, 'html': render_to_string(self.template_name, {'form': form}, request=self.request)})
        else:
            return super().form_invalid(form)


# --------------------------- Formulario Crear  Detalle De Pedido ----------------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class DetallePedidoCreateView(CreateView):
    model = DetallePedido
    form_class = DetallePedidoForm
    template_name = 'detalle_pedido_crear.html'
    success_url = '/detalle_pedido_crear/'

    def get_initial(self):
        initial = super().get_initial()
        pedido_id = self.kwargs.get('pedido_id') or self.request.GET.get('pedido_id')
        if pedido_id:
            initial['pedido'] = pedido_id
        return initial

    def get_form_kwargs(self):
        kwargs = super(DetallePedidoCreateView, self).get_form_kwargs()
        kwargs['pedido_id'] = self.kwargs.get('pedido_id')
        return kwargs

    @transaction.atomic
    def form_valid(self, form):
        pedido_id = self.kwargs.get('pedido_id')
        if pedido_id:
            pedido = get_object_or_404(Pedido, pk=pedido_id)
            form.instance.pedido = pedido

        self.object = form.save()
        messages.success(self.request, f'El detalle de pedido para el pedido {pedido_id} se ha creado exitosamente.')
        return JsonResponse({'success': True})

    def form_invalid(self, form):
        errors = form.errors.as_json()
        return JsonResponse({'success': False, 'error': errors})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pedido_id'] = self.kwargs.get('pedido_id')
        return context


# ---------------------------- Formulario Editar Detalle De Pedido --------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class DetallePedidoUpdateView(UpdateView):
    model = DetallePedido
    form_class = EditarDetallePedidoForm
    template_name = 'detalle_pedido_editar.html'
    success_url = '/detalle_pedido_editar/'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.object = None

    def get_object(self, queryset=None):
        detallepedido_id = self.request.POST.get('detallepedido_id')
        detallepedido = get_object_or_404(DetallePedido, id=detallepedido_id)
        return detallepedido

    def get(self, request, *args, **kwargs):
        detallepedido_id = request.GET.get('detallepedido_id')
        self.object = get_object_or_404(DetallePedido, id=detallepedido_id)
        form = self.form_class(
            instance=self.object,
        )
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, {'form': form}, request=request)
            return JsonResponse({'form': form_html})
        else:
            return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request,
                         f'El detalle para el {form.cleaned_data['pedido']} se ha editado exitosamente.')
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse(
                {'success': False, 'html': render_to_string(self.template_name, {'form': form}, request=self.request)})
        else:
            return super().form_invalid(form)


# ---------------------------- Formulario Eliminar Detalle De Pedido --------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class DetallePedidoDeleteiew(UpdateView):
    model = DetallePedido
    form_class = EliminarDetallePedidoForm
    template_name = 'detalle_pedido_eliminart.html'
    success_url = '/detalle_pedido_eliminar/'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.object = None

    def get_object(self, queryset=None):
        detallepedido_id = self.request.POST.get('detallepedido_id')
        detallepedido = get_object_or_404(DetallePedido, id=detallepedido_id)
        return detallepedido

    def get(self, request, *args, **kwargs):
        detallepedido_id = request.GET.get('detallepedido_id')
        self.object = get_object_or_404(DetallePedido, id=detallepedido_id)
        form = self.form_class(
            instance=self.object,
        )
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, {'form': form}, request=request)
            return JsonResponse({'form': form_html})
        else:
            return super().get(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['instance'] = self.object
        return kwargs

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        detallepedido = form.save(commit=False)
        detallepedido.delete()
        messages.success(self.request,
                         f'El detalle de {form.cleaned_data['pedido']} se ha eliminado exitosamente.')
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse(
                {'success': False, 'html': render_to_string(self.template_name, {'form': form}, request=self.request)})
        else:
            return super().form_invalid(form)


# ------------------------- CARTERA General /// Table mostrar cartera de pedidos Heavens ///  -------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class CarteraHeavensListView(SingleTableView):
    model = Pedido
    table_class = CarteraPedidoTable
    table_pagination = {"per_page": 14}
    template_name = 'cartera_list_heavens.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset()
        form = self.form_class(self.request.GET)

        if form.is_valid() and form.cleaned_data.get('item_busqueda'):
            item_busqueda = form.cleaned_data.get('item_busqueda')

            # Aquí se maneja la lógica de la búsqueda
            try:
                item_busqueda_id = int(item_busqueda)  # Intenta convertir a entero
                queryset = queryset.filter(Q(cliente__nombre__icontains=item_busqueda) | Q(id=item_busqueda_id))
            except ValueError:
                # Si la conversión falla, busca solo por nombre
                queryset = queryset.filter(cliente__nombre__icontains=item_busqueda)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context


# ------------------------- CARTERA /// Table mostrar cartera de pedidos Etnico ///  ---------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Etnico'), login_url=reverse_lazy('home')), name='dispatch')
class CarteraEtnicoListView(SingleTableView):
    model = Pedido
    table_class = CarteraPedidoTable
    table_pagination = {"per_page": 14}
    template_name = 'cartera_list_etnico.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Etnico')
        form = self.form_class(self.request.GET)

        if form.is_valid() and form.cleaned_data.get('item_busqueda'):
            item_busqueda = form.cleaned_data.get('item_busqueda')

            # Aquí se maneja la lógica de la búsqueda
            try:
                item_busqueda_id = int(item_busqueda)  # Intenta convertir a entero
                queryset = queryset.filter(Q(cliente__nombre__icontains=item_busqueda) | Q(id=item_busqueda_id))
            except ValueError:
                # Si la conversión falla, busca solo por nombre
                queryset = queryset.filter(cliente__nombre__icontains=item_busqueda)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context


# ------------------------- CARTERA /// Table mostrar cartera de pedidos Fieldex ///  ---------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Fieldex'), login_url=reverse_lazy('home')), name='dispatch')
class CarteraFieldexListView(SingleTableView):
    model = Pedido
    table_class = CarteraPedidoTable
    table_pagination = {"per_page": 14}
    template_name = 'cartera_list_fieldex.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Fieldex')
        form = self.form_class(self.request.GET)

        if form.is_valid() and form.cleaned_data.get('item_busqueda'):
            item_busqueda = form.cleaned_data.get('item_busqueda')

            # Aquí se maneja la lógica de la búsqueda
            try:
                item_busqueda_id = int(item_busqueda)  # Intenta convertir a entero
                queryset = queryset.filter(Q(cliente__nombre__icontains=item_busqueda) | Q(id=item_busqueda_id))
            except ValueError:
                # Si la conversión falla, busca solo por nombre
                queryset = queryset.filter(cliente__nombre__icontains=item_busqueda)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context


# ------------------------- CARTERA /// Table mostrar cartera de pedidos Juan Matas ///  ---------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Juan_Matas'), login_url=reverse_lazy('home')), name='dispatch')
class CarteraJuanListView(SingleTableView):
    model = Pedido
    table_class = CarteraPedidoTable
    table_pagination = {"per_page": 14}
    template_name = 'cartera_list_juan.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Juan_Matas')
        form = self.form_class(self.request.GET)

        if form.is_valid() and form.cleaned_data.get('item_busqueda'):
            item_busqueda = form.cleaned_data.get('item_busqueda')

            # Aquí se maneja la lógica de la búsqueda
            try:
                item_busqueda_id = int(item_busqueda)  # Intenta convertir a entero
                queryset = queryset.filter(Q(cliente__nombre__icontains=item_busqueda) | Q(id=item_busqueda_id))
            except ValueError:
                # Si la conversión falla, busca solo por nombre
                queryset = queryset.filter(cliente__nombre__icontains=item_busqueda)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context


# ------------------------- COMISIONES GENERAL /// Table mostrar comisiones de pedidos Heavens ///  ------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class ComisionHeavensListView(SingleTableView):
    model = Pedido
    table_class = ComisionPedidoTable
    table_pagination = {"per_page": 14}
    template_name = 'comision_list_heavens.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset()
        form = self.form_class(self.request.GET)

        if form.is_valid() and form.cleaned_data.get('item_busqueda'):
            item_busqueda = form.cleaned_data.get('item_busqueda')

            # Aquí se maneja la lógica de la búsqueda
            try:
                item_busqueda_id = int(item_busqueda)  # Intenta convertir a entero
                queryset = queryset.filter(Q(cliente__nombre__icontains=item_busqueda) | Q(id=item_busqueda_id))
            except ValueError:
                # Si la conversión falla, busca solo por nombre
                queryset = queryset.filter(cliente__nombre__icontains=item_busqueda)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context


# ------------------------- COMISIONES /// Table mostrar comisiones de pedidos Etnico ///  ---------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Etnico'), login_url=reverse_lazy('home')), name='dispatch')
class ComisionEtnicoListView(SingleTableView):
    model = Pedido
    table_class = ComisionPedidoTable
    table_pagination = {"per_page": 14}
    template_name = 'comision_list_etnico.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Etnico')
        form = self.form_class(self.request.GET)

        if form.is_valid() and form.cleaned_data.get('item_busqueda'):
            item_busqueda = form.cleaned_data.get('item_busqueda')

            # Aquí se maneja la lógica de la búsqueda
            try:
                item_busqueda_id = int(item_busqueda)  # Intenta convertir a entero
                queryset = queryset.filter(Q(cliente__nombre__icontains=item_busqueda) | Q(id=item_busqueda_id))
            except ValueError:
                # Si la conversión falla, busca solo por nombre
                queryset = queryset.filter(cliente__nombre__icontains=item_busqueda)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context


# ------------------------- COMISIONES /// Table mostrar comisiones de pedidos Fieldex ///  ---------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Fieldex'), login_url=reverse_lazy('home')), name='dispatch')
class ComisionFiedexListView(SingleTableView):
    model = Pedido
    table_class = ComisionPedidoTable
    table_pagination = {"per_page": 14}
    template_name = 'comision_list_fieldex.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Fieldex')
        form = self.form_class(self.request.GET)

        if form.is_valid() and form.cleaned_data.get('item_busqueda'):
            item_busqueda = form.cleaned_data.get('item_busqueda')

            # Aquí se maneja la lógica de la búsqueda
            try:
                item_busqueda_id = int(item_busqueda)  # Intenta convertir a entero
                queryset = queryset.filter(Q(cliente__nombre__icontains=item_busqueda) | Q(id=item_busqueda_id))
            except ValueError:
                # Si la conversión falla, busca solo por nombre
                queryset = queryset.filter(cliente__nombre__icontains=item_busqueda)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context


# ------------------------- COMISIONES /// Table mostrar comisiones de pedidos Juan Matas // -----------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Juan_Matas'), login_url=reverse_lazy('home')), name='dispatch')
class ComisionJuanListView(SingleTableView):
    model = Pedido
    table_class = ComisionPedidoTable
    table_pagination = {"per_page": 14}
    template_name = 'comision_list_juan.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Juan_Matas')
        form = self.form_class(self.request.GET)

        if form.is_valid() and form.cleaned_data.get('item_busqueda'):
            item_busqueda = form.cleaned_data.get('item_busqueda')

            # Aquí se maneja la lógica de la búsqueda
            try:
                item_busqueda_id = int(item_busqueda)  # Intenta convertir a entero
                queryset = queryset.filter(Q(cliente__nombre__icontains=item_busqueda) | Q(id=item_busqueda_id))
            except ValueError:
                # Si la conversión falla, busca solo por nombre
                queryset = queryset.filter(cliente__nombre__icontains=item_busqueda)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context
